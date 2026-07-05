"""Excel planner parser — the single "front door" from spreadsheet to profile dict.

`parse_planner(path)` turns either provided planner (Arjun / Priya) into ONE
internal profile dict that every downstream component uses (finance_tools,
finance_agent, redactor, profile_store).

Two hard lessons baked into this parser, learned by inspecting the ACTUAL files:

1. The two planners do NOT share a row layout. Arjun has 9 expense rows and
   Priya has 8, which pushes every section below Expenses onto different row
   numbers (Arjun's Net Worth is B35, Priya's is B34). So we NEVER hardcode row
   numbers — we locate each section by its LABEL in column A/D and read the
   rows between a section header and its "Total" row. This survives small
   layout edits to either planner.

2. Labels differ too ("Home Loan / Rent EMI" vs "Rent", "Spouse Net Income" vs
   "Freelance Income"). So income/expense/asset line items are stored under
   normalized-from-the-sheet keys, not a fixed vocabulary. finance_tools only
   ever need the *totals*, which it recomputes — it never trusts the sheet's
   own Total cells.

Edge cases handled without error (Priya): empty debts (all liabilities zero ->
`debts == []`) and zero insurance (`life_cover == 0`, `health_cover == 0`).
"""
from __future__ import annotations

import os
import re
from typing import Any

import openpyxl

# --- Representative debt interest rates -------------------------------------
# The planner's Liabilities section captures only OUTSTANDING BALANCES — it has
# no interest-rate or EMI column. But the debt-avalanche / high-interest logic
# genuinely needs rates. These are REPRESENTATIVE annual rates (%) for the
# Indian market by debt type, applied by the parser and clearly surfaced as
# assumptions, not sourced from the sheet. The ~42% credit-card figure is the
# well-known typical Indian card APR (~3.5%/month) and is what drives the
# "clear the 42% card first" advice in the demo.
REPRESENTATIVE_ANNUAL_RATES = {
    "home_loan": 8.5,
    "car_loan": 10.0,
    "credit_card": 42.0,
    "personal_loan": 16.0,
    "other_loan": 12.0,
}


def _norm_key(label: str) -> str:
    """'Spouse Net Income' -> 'spouse_net_income'. Stable snake_case key."""
    label = label.strip().lower()
    label = re.sub(r"[^a-z0-9]+", "_", label)
    return label.strip("_")


def _num(value: Any) -> float:
    """Coerce a cell to a non-negative float; blank/None -> 0.0.

    Guardrail from the spec: numbers are non-negative and missing optional
    cells default to 0. A genuinely non-numeric value where a number is
    required raises, so junk never silently flows to the model.
    """
    if value is None or value == "":
        return 0.0
    try:
        n = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Expected a number but got {value!r}") from exc
    return max(0.0, n)


def _build_colmap(ws) -> dict[str, list[tuple[int, str]]]:
    """Map normalized column-A and column-D text -> list of (row, raw_label).

    We keep a list per key because a sheet can repeat labels; callers pick by
    section boundaries.
    """
    out: dict[str, list[tuple[int, str]]] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.column_letter in ("A", "D") and isinstance(cell.value, str):
                key = _norm_key(cell.value)
                out.setdefault(key, []).append((cell.row, cell.value))
    return out


def _find_row(ws, label_col: str, contains: str, exact: bool = False) -> int | None:
    """First row whose `label_col` cell matches `contains` (case-insensitive).

    `exact=True` requires the whole cell to equal `contains` (after strip) — use
    it when a substring would collide with another label, e.g. the "Goal" table
    header vs the "Financial Goals" sheet title.
    """
    needle = contains.lower()
    for row in ws.iter_rows():
        for cell in row:
            if cell.column_letter == label_col and isinstance(cell.value, str):
                text = cell.value.strip().lower()
                if (text == needle) if exact else (needle in text):
                    return cell.row
    return None


def _section_items(
    ws, header_contains: str, total_contains: str, label_col: str, value_col: str
) -> dict[str, float]:
    """Read (label -> value) pairs strictly between a header row and its Total row.

    e.g. everything between the "Monthly Income" header and "Total Income".
    Matching by these two anchors is what makes the parser layout-independent.
    """
    start = _find_row(ws, label_col, header_contains)
    end = _find_row(ws, label_col, total_contains)
    if start is None or end is None:
        raise ValueError(
            f"Could not locate section '{header_contains}'..'{total_contains}' "
            f"in sheet '{ws.title}'"
        )
    items: dict[str, float] = {}
    for r in range(start + 1, end):
        label = ws[f"{label_col}{r}"].value
        if not isinstance(label, str) or not label.strip():
            continue
        items[_norm_key(label)] = _num(ws[f"{value_col}{r}"].value)
    return items


def _labeled_value(ws, contains: str, label_col: str = "A", value_col: str = "B") -> float:
    """Value next to the first label containing `contains`. 0.0 if absent."""
    row = _find_row(ws, label_col, contains)
    if row is None:
        return 0.0
    return _num(ws[f"{value_col}{row}"].value)


def _name_from_path(path: str) -> str:
    """Derive the display name from the filename (…Profile1_Arjun.xlsx -> 'Arjun').

    The planner template has no structured name cell, so the filename is the
    most reliable source. The privacy layer will redact this anyway.
    """
    stem = os.path.splitext(os.path.basename(path))[0]
    parts = stem.split("_")
    return parts[-1] if parts else stem


def _infer_dependents(goals: list[dict]) -> int:
    """Infer child dependents from unique 'Child N' tokens in goal names.

    Not captured by the template; inferred so insurance/HLV framing is sensible.
    Arjun's goals reference Child 1 and Child 2 -> 2; Priya's reference none -> 0.
    """
    children = set()
    for g in goals:
        for match in re.findall(r"child\s*(\d+)", g["name"].lower()):
            children.add(match)
    return len(children)


def _infer_risk_appetite(age: int) -> str:
    """Heuristic (age-based), since the template captures no risk field.

    Documented as a heuristic: younger investors can typically ride more equity
    volatility. 26 -> aggressive (Priya), 36 -> moderate (Arjun).
    """
    if age < 30:
        return "aggressive"
    if age <= 45:
        return "moderate"
    return "conservative"


def parse_planner(path: str) -> dict:
    """Parse a PB planner .xlsx into the internal profile dict.

    Reads with data_only=True so formula cells arrive as their computed values.
    Recomputation of all totals is deliberately left to finance_tools — this
    parser only extracts inputs.
    """
    if not os.path.exists(path):
        raise FileNotFoundError(f"Planner file not found: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    required = {"CashFlow_NetWorth", "Retirement", "Goals", "HLV"}
    missing = required - set(wb.sheetnames)
    if missing:
        raise ValueError(f"Planner is missing required sheet(s): {sorted(missing)}")

    cf = wb["CashFlow_NetWorth"]
    ret = wb["Retirement"]
    goals_ws = wb["Goals"]
    hlv = wb["HLV"]

    # --- Cash flow: income / expenses (label-anchored sections) ---
    income = _section_items(cf, "Monthly Income", "Total Income", "A", "B")
    expenses = _section_items(cf, "Monthly Expenses", "Total Expenses", "A", "B")

    # --- Assets (col A/B) between header and total ---
    assets = _section_items(cf, "Assets (current value)", "Total Assets", "A", "B")

    # --- Liabilities (col D/E) -> debts list; zero balances skipped ---
    liab_items = _section_items(cf, "Liabilities", "Total Liabilities", "D", "E")
    debts: list[dict] = []
    for key, balance in liab_items.items():
        if balance <= 0:  # skip zero-balance liabilities -> Priya ends up with []
            continue
        rate = REPRESENTATIVE_ANNUAL_RATES.get(key)
        debts.append(
            {
                "type": key,
                "balance": balance,
                # rate is a representative assumption (see module docstring);
                # None only if we have no rate mapping for this label.
                "rate_pct": rate,
                "rate_is_assumed": rate is not None,
                "emi": None,  # not captured by the template
            }
        )

    # --- Insurance (zero is valid — Priya) ---
    insurance = {
        "life_cover": _labeled_value(cf, "Life Insurance"),
        "health_cover": _labeled_value(cf, "Health Insurance"),
    }

    # --- Goals table: rows between the "Goal" header and the "TOTAL" row ---
    goals: list[dict] = []
    header_row = _find_row(goals_ws, "A", "Goal", exact=True)
    if header_row is not None:
        r = header_row + 1
        while True:
            name = goals_ws[f"A{r}"].value
            if not isinstance(name, str) or not name.strip():
                break
            if name.strip().upper().startswith("TOTAL"):
                break
            goals.append(
                {
                    "name": name.strip(),
                    "cost": _num(goals_ws[f"B{r}"].value),
                    "years": int(_num(goals_ws[f"C{r}"].value)),
                    "inflation": _num(goals_ws[f"D{r}"].value),
                }
            )
            r += 1

    # --- Retirement inputs (label-matched; calculators recompute outputs) ---
    retirement_inputs = {
        "current_age": int(_labeled_value(ret, "Current Age")),
        "retirement_age": int(_labeled_value(ret, "Retirement Age")),
        "life_expectancy": int(_labeled_value(ret, "Life Expectancy")),
        "monthly_expense": _labeled_value(ret, "Current Monthly Expenses"),
        "inflation": _labeled_value(ret, "Inflation"),
        "roi_pre": _labeled_value(ret, "pre-retirement"),
        # "Return post-retirement" — anchor on "Return post" so we don't collide
        # with "Current Monthly Expenses (post-retirement)".
        "roi_post": _labeled_value(ret, "Return post"),
        "already_invested": _labeled_value(ret, "Already Invested"),
    }

    # --- HLV inputs (assumptions specific to this sheet) ---
    hlv_inputs = {
        "annual_income_to_replace": _labeled_value(hlv, "Annual Income to Replace"),
        "years_to_provide": int(_labeled_value(hlv, "Years to Provide")),
        "inflation": _labeled_value(hlv, "Inflation"),
        "roi": _labeled_value(hlv, "Return on Investment"),
    }

    age = retirement_inputs["current_age"]

    profile = {
        "meta": {"source_file": os.path.abspath(path)},
        "profile": {
            "name": _name_from_path(path),
            "age": age,
            "retirement_age": retirement_inputs["retirement_age"],
            "life_expectancy": retirement_inputs["life_expectancy"],
            "dependents": _infer_dependents(goals),
            "risk_appetite": _infer_risk_appetite(age),
            "city": None,  # not captured by the template
        },
        "income_monthly": income,
        "expenses_monthly": expenses,
        "assets": assets,
        "debts": debts,
        "insurance": insurance,
        "goals": goals,
        "retirement_inputs": retirement_inputs,
        "hlv_inputs": hlv_inputs,
        # filled in by finance_tools.compute_derived(); kept here so the schema
        # is stable and self-documenting even before calculators run.
        "derived": {},
    }
    return profile
